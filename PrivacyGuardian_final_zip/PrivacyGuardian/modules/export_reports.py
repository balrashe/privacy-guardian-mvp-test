"""
Export Reports module for Privacy Guardian.

Generates compliance reports and audit trails in PDF and Excel formats
for regulatory documentation and business records.
"""
from __future__ import annotations
import io
import pandas as pd
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Union
import base64

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ComplianceReportGenerator:
    """Generates comprehensive compliance reports in multiple formats."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Set up custom styles for PDF generation."""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50')
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=12,
            spaceAfter=6,
            textColor=colors.HexColor('#34495e')
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            alignment=TA_JUSTIFY
        ))

    def generate_risk_assessment_report_pdf(
        self, 
        assessment_data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> bytes:
        """Generate a comprehensive risk assessment report in PDF format."""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72, leftMargin=72,
            topMargin=72, bottomMargin=18
        )
        
        story = []
        
        # Title
        title = f"Data Risk Assessment Report"
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # Report metadata
        metadata = [
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Dataset:', assessment_data.get('dataset_name', 'Unknown')],
            ['Classification Method:', assessment_data.get('method', 'Rule-based')],
            ['Total Columns:', str(assessment_data.get('total_columns', 0))],
            ['Total Rows:', str(assessment_data.get('total_rows', 0))]
        ]
        
        metadata_table = Table(metadata, colWidths=[2*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 20))
        
        # Risk Summary
        story.append(Paragraph("Risk Level Summary", self.styles['CustomHeading']))
        summary = assessment_data.get('risk_summary', {})
        
        summary_data = [['Risk Level', 'Count', 'Percentage']]
        total = sum(summary.values()) if summary else 1
        for level in ['High', 'Medium', 'Low']:
            count = summary.get(level, 0)
            percentage = f"{(count/total)*100:.1f}%" if total > 0 else "0%"
            summary_data.append([level, str(count), percentage])
        
        summary_table = Table(summary_data, colWidths=[1.5*inch, 1*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Detailed Classification Results
        story.append(Paragraph("Detailed Classification Results", self.styles['CustomHeading']))
        
        results = assessment_data.get('classification_results', [])
        if results:
            # Prepare table data
            if 'hybrid_final_risk' in results[0]:
                # Hybrid results
                table_data = [['Column Name', 'Final Risk', 'Method', 'Confidence']]
                for result in results:
                    table_data.append([
                        result.get('column', ''),
                        result.get('hybrid_final_risk', ''),
                        result.get('hybrid_method', ''),
                        f"{result.get('confidence_score', 0):.3f}"
                    ])
            else:
                # Rule-based results
                table_data = [['Column Name', 'Name Risk', 'Value Risk', 'Final Risk']]
                for result in results:
                    table_data.append([
                        result.get('column', ''),
                        result.get('name_hint_risk', ''),
                        result.get('value_sample_risk', ''),
                        result.get('final_risk', '')
                    ])
            
            results_table = Table(table_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(results_table)
        
        # Recommendations
        story.append(PageBreak())
        story.append(Paragraph("Recommendations", self.styles['CustomHeading']))
        
        recommendations = [
            "• Review columns classified as 'High' risk for additional security measures",
            "• Consider data minimization for unnecessary personal information",
            "• Implement appropriate technical and organizational measures",
            "• Document data processing purposes and lawful bases",
            "• Establish data retention and deletion schedules",
            "• Conduct regular privacy impact assessments"
        ]
        
        for rec in recommendations:
            story.append(Paragraph(rec, self.styles['CustomBody']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_compliance_audit_report_pdf(
        self,
        audit_data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> bytes:
        """Generate a compliance audit report in PDF format."""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72, leftMargin=72,
            topMargin=72, bottomMargin=18
        )
        
        story = []
        
        # Title
        title = "Privacy Compliance Audit Report"
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # Audit metadata
        metadata = [
            ['Audit Date:', datetime.now().strftime('%Y-%m-%d')],
            ['Organization:', audit_data.get('organization', 'N/A')],
            ['Scope:', audit_data.get('scope', 'Privacy Management Program')],
            ['Auditor:', audit_data.get('auditor', 'Privacy Guardian System')],
            ['Compliance Score:', f"{audit_data.get('score', 0)}/{audit_data.get('max_score', 0)} ({audit_data.get('percentage', 0):.0f}%)"]
        ]
        
        metadata_table = Table(metadata, colWidths=[2*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 20))
        
        # Compliance Results
        story.append(Paragraph("Compliance Assessment Results", self.styles['CustomHeading']))
        
        responses = audit_data.get('responses', {})
        checklist = audit_data.get('checklist', {})
        
        table_data = [['Control Area', 'Status', 'Weight', 'Question']]
        for key, response in responses.items():
            question_data = checklist.get(key, {})
            status = "✓ Yes" if response.lower() == 'yes' else "✗ No"
            table_data.append([
                key.replace('_', ' ').title(),
                status,
                str(question_data.get('weight', 1)),
                question_data.get('question', '')[:60] + '...' if len(question_data.get('question', '')) > 60 else question_data.get('question', '')
            ])
        
        results_table = Table(table_data, colWidths=[1.5*inch, 0.8*inch, 0.7*inch, 3*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        story.append(results_table)
        story.append(Spacer(1, 20))
        
        # Recommendations
        story.append(Paragraph("Priority Recommendations", self.styles['CustomHeading']))
        recommendations = audit_data.get('recommendations', [])
        
        if recommendations:
            for i, rec in enumerate(recommendations, 1):
                story.append(Paragraph(f"{i}. {rec}", self.styles['CustomBody']))
        else:
            story.append(Paragraph("No specific recommendations - compliance score indicates good privacy practices.", self.styles['CustomBody']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel_report(
        self,
        assessment_data: Dict[str, Any],
        audit_data: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate a comprehensive Excel report with multiple worksheets."""
        
        buffer = io.BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Data styles
        data_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Risk Assessment Sheet
        ws_risk = wb.create_sheet("Risk Assessment")
        
        # Headers
        headers = ["Column Name", "Name Hint Risk", "Value Sample Risk", "Final Risk"]
        if assessment_data.get('classification_results') and 'hybrid_final_risk' in assessment_data['classification_results'][0]:
            headers = ["Column Name", "Hybrid Final Risk", "Hybrid Method", "Confidence Score", "ML Name Risk", "ML Data Risk"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_risk.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        results = assessment_data.get('classification_results', [])
        for row, result in enumerate(results, 2):
            if 'hybrid_final_risk' in result:
                # Hybrid data
                data = [
                    result.get('column', ''),
                    result.get('hybrid_final_risk', ''),
                    result.get('hybrid_method', ''),
                    result.get('confidence_score', ''),
                    result.get('ml_name_risk', ''),
                    result.get('ml_data_risk', '')
                ]
            else:
                # Rule-based data
                data = [
                    result.get('column', ''),
                    result.get('name_hint_risk', ''),
                    result.get('value_sample_risk', ''),
                    result.get('final_risk', '')
                ]
            
            for col, value in enumerate(data, 1):
                cell = ws_risk.cell(row=row, column=col, value=value)
                cell.alignment = data_alignment
                cell.border = border
                
                # Color code by risk level
                if isinstance(value, str) and value in ['High', 'Medium', 'Low']:
                    if value == 'High':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value == 'Medium':
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                    elif value == 'Low':
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_risk.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_risk.column_dimensions[column_letter].width = adjusted_width
        
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Risk summary
        ws_summary.cell(row=1, column=1, value="Risk Level Summary").font = Font(bold=True, size=14)
        
        summary_headers = ["Risk Level", "Count", "Percentage"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        summary = assessment_data.get('risk_summary', {})
        total = sum(summary.values()) if summary else 1
        
        for row, level in enumerate(['High', 'Medium', 'Low'], 4):
            count = summary.get(level, 0)
            percentage = f"{(count/total)*100:.1f}%" if total > 0 else "0%"
            
            ws_summary.cell(row=row, column=1, value=level).border = border
            ws_summary.cell(row=row, column=2, value=count).border = border
            ws_summary.cell(row=row, column=3, value=percentage).border = border
        
        # Compliance Audit Sheet (if data provided)
        if audit_data:
            ws_audit = wb.create_sheet("Compliance Audit")
            
            # Audit headers
            audit_headers = ["Control Area", "Question", "Response", "Weight", "Recommendation"]
            for col, header in enumerate(audit_headers, 1):
                cell = ws_audit.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Audit data
            responses = audit_data.get('responses', {})
            checklist = audit_data.get('checklist', {})
            recommendations = audit_data.get('recommendations', [])
            
            rec_index = 0
            for row, (key, response) in enumerate(responses.items(), 2):
                question_data = checklist.get(key, {})
                recommendation = recommendations[rec_index] if rec_index < len(recommendations) and response.lower() == 'no' else ""
                if response.lower() == 'no' and rec_index < len(recommendations):
                    rec_index += 1
                
                data = [
                    key.replace('_', ' ').title(),
                    question_data.get('question', ''),
                    response,
                    question_data.get('weight', 1),
                    recommendation
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws_audit.cell(row=row, column=col, value=value)
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # Color code responses
                    if col == 3:  # Response column
                        if response.lower() == 'yes':
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws_audit.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_audit.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def prepare_assessment_data(
    classification_results: List[Dict[str, Any]],
    risk_summary: Dict[str, int],
    dataset_name: str = "Unknown Dataset",
    total_rows: int = 0,
    method: str = "Rule-based"
) -> Dict[str, Any]:
    """Prepare assessment data for report generation."""
    return {
        'classification_results': classification_results,
        'risk_summary': risk_summary,
        'dataset_name': dataset_name or "Unknown Dataset",
        'total_columns': len(classification_results) if classification_results else 0,
        'total_rows': total_rows if total_rows > 0 else 0,
        'method': method or "Rule-based"
    }


def prepare_audit_data(
    responses: Dict[str, str],
    checklist: Dict[str, Any],
    score: int,
    max_score: int,
    recommendations: List[str],
    organization: str = "Organization"
) -> Dict[str, Any]:
    """Prepare audit data for report generation."""
    return {
        'responses': responses,
        'checklist': checklist,
        'score': score,
        'max_score': max_score,
        'percentage': (score / max_score) * 100 if max_score > 0 else 0,
        'recommendations': recommendations,
        'organization': organization,
        'scope': 'Privacy Management Program',
        'auditor': 'Privacy Guardian System'
    }


# Convenience functions for easy integration
def export_risk_assessment_pdf(
    classification_results: List[Dict[str, Any]],
    risk_summary: Dict[str, int],
    dataset_name: str = "Unknown Dataset",
    total_rows: int = 0,
    method: str = "Rule-based"
) -> bytes:
    """Export risk assessment as PDF."""
    generator = ComplianceReportGenerator()
    data = prepare_assessment_data(classification_results, risk_summary, dataset_name, total_rows, method)
    return generator.generate_risk_assessment_report_pdf(data)


def export_compliance_audit_pdf(
    responses: Dict[str, str],
    checklist: Dict[str, Any],
    score: int,
    max_score: int,
    recommendations: List[str],
    organization: str = "Organization"
) -> bytes:
    """Export compliance audit as PDF."""
    generator = ComplianceReportGenerator()
    data = prepare_audit_data(responses, checklist, score, max_score, recommendations, organization)
    return generator.generate_compliance_audit_report_pdf(data)


def export_combined_excel_report(
    classification_results: List[Dict[str, Any]],
    risk_summary: Dict[str, int],
    responses: Optional[Dict[str, str]] = None,
    checklist: Optional[Dict[str, Any]] = None,
    score: Optional[int] = None,
    max_score: Optional[int] = None,
    recommendations: Optional[List[str]] = None,
    dataset_name: str = "Unknown Dataset",
    total_rows: int = 0,
    method: str = "Rule-based",
    organization: str = "Organization"
) -> bytes:
    """Export combined assessment and audit report as Excel."""
    generator = ComplianceReportGenerator()
    
    assessment_data = prepare_assessment_data(classification_results, risk_summary, dataset_name, total_rows, method)
    
    audit_data = None
    if responses and checklist and score is not None and max_score is not None:
        audit_data = prepare_audit_data(responses, checklist, score, max_score, recommendations or [], organization)
    
    return generator.generate_excel_report(assessment_data, audit_data)